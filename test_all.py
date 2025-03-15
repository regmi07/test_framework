import os
import json
import importlib.util
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook, load_workbook
import logging
import ast
import subprocess
import platform
from collections import defaultdict

import inspect

# configure logging
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")


def load_solution_function(solution_path, function_name, language):
    """
    Dynamically import or compile the solution and return a runner function.
    """
    try:
        parts = solution_path.split(os.path.sep)
        username = parts[-3]  # Extract username from path structure
        print(solution_path)

        if language == 'python':
            print('inside here')
            # Load Python function
            spec = importlib.util.spec_from_file_location(
                f"Solution_{username}", solution_path)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            func = getattr(module, function_name, None)
            print(inspect.isfunction(func))
            return (func, username, language) if func else (None, username, language)
        else:
            solution_dir = os.path.dirname(solution_path)
            if language == 'java':
                # Compile Java solution
                compile_cmd = ['javac', solution_path]
                try:
                    subprocess.run(compile_cmd, check=True,
                                   capture_output=True)
                except subprocess.CalledProcessError as e:
                    logging.error(
                        f"Java compilation failed for {solution_path}: {e.stderr.decode()}")
                    return (None, username, language)

                def java_runner(inputs):
                    args = [str(i) for i in (inputs if isinstance(
                        inputs, (list, tuple)) else [inputs])]
                    print('args: ', args)
                    cmd = ['java', '-cp', solution_dir, 'Solution'] + args
                    try:
                        result = subprocess.run(
                            cmd, capture_output=True, text=True, timeout=5)
                        return result.stdout.strip()
                    except Exception as e:
                        logging.error(f"Error running Java solution: {e}")
                        return None
                return (java_runner, username, language)

            elif language == 'c':
                # Compile C solution
                executable = os.path.join(
                    solution_dir, 'solution.exe' if platform.system() == 'Windows' else 'solution')
                compile_cmd = ['gcc', solution_path, '-o', executable]
                try:
                    subprocess.run(compile_cmd, check=True,
                                   capture_output=True)
                except subprocess.CalledProcessError as e:
                    logging.error(
                        f"C compilation failed for {solution_path}: {e.stderr.decode()}")
                    return (None, username, language)

                def c_runner(inputs):
                    args = [str(i) for i in (inputs if isinstance(
                        inputs, (list, tuple)) else [inputs])]
                    print('args: ', args)
                    cmd = [executable] + args
                    try:
                        result = subprocess.run(
                            cmd, capture_output=True, text=True, timeout=5)
                        return result.stdout.strip()
                    except Exception as e:
                        logging.error(f"Error running C solution: {e}")
                        return None
                return (c_runner, username, language)

            elif language == 'cpp':
                # Compile C++ solution
                executable = os.path.join(
                    solution_dir, 'solution.exe' if platform.system() == 'Windows' else 'solution')
                compile_cmd = ['g++', solution_path, '-o', executable]
                try:
                    subprocess.run(compile_cmd, check=True,
                                   capture_output=True)
                except subprocess.CalledProcessError as e:
                    logging.error(
                        f"C++ compilation failed for {solution_path}: {e.stderr.decode()}")
                    return (None, username, language)

                def cpp_runner(inputs):
                    args = [str(i) for i in (inputs if isinstance(
                        inputs, (list, tuple)) else [inputs])]
                    print('args: ', args)
                    cmd = [executable] + args
                    try:
                        result = subprocess.run(
                            cmd, capture_output=True, text=True, timeout=5)
                        return result.stdout.strip()
                    except Exception as e:
                        logging.error(f"Error running C++ solution: {e}")
                        return None
                return (cpp_runner, username, language)

            elif language == 'js':
                # Node.js solution runner
                def js_runner(inputs):
                    args = [str(i) for i in (inputs if isinstance(
                        inputs, (list, tuple)) else [inputs])]
                    print('args: ', args)
                    cmd = ['node', solution_path] + args
                    try:
                        result = subprocess.run(
                            cmd, capture_output=True, text=True, timeout=5)
                        return result.stdout.strip()
                    except Exception as e:
                        logging.error(f"Error running JS solution: {e}")
                        return None
                return (js_runner, username, language)

            else:
                logging.error(f"Unsupported language: {language}")
                return (None, username, language)

    except Exception as e:
        logging.error(f"Failed to load {solution_path}: {e}")
        return (None, None, language)


def run_test_for_user(runner, test_cases, language):
    """
    Run test cases for a solution and return passed tests.
    """
    passed = 0
    for inputs, expected in test_cases:
        try:
            if language == 'python':
                result = runner(*inputs) if isinstance(inputs,
                                                       (list, tuple)) else runner(inputs)
                if result == expected:
                    passed += 1
            else:
                print('inputs: ', inputs)
                output = runner(inputs)
                print('out: ', output, ' expected: ', expected)

                if output is None:
                    continue
                # Try to parse output and compare with expected
                try:
                    parsed = ast.literal_eval(output)
                    if parsed == expected:
                        passed += 1
                    elif str(parsed) == str(expected):
                        passed += 1
                except:
                    if output.strip().lower() == str(expected).strip().lower():
                        passed += 1
        except Exception as e:
            logging.error(f"Test execution error: {e}")
    return passed


def calculate_score(passed, total, max_score):
    return (passed / total) * max_score if total else 0


def run_test_for_challenge(challenge, paths, function_name, test_cases, max_score):
    """
    Process all solutions for a challenge and return results.
    """
    results = []
    for solution_path, language in paths:
        runner, username, _ = load_solution_function(
            solution_path, function_name, language)
        if not runner or not username:
            print('i am not running')
            continue
        passed = run_test_for_user(runner, test_cases, language)
        score = calculate_score(passed, len(test_cases), max_score)
        results.append((username, challenge, score))
    return results


def update_excel(results, excel_file="code_clash_result2.xlsx"):
    """
        Update an Excel sheet with usernames and their scores.
    """
    user_scores = {}
    for username, challenge, score in results:
        if username not in user_scores:
            user_scores[username] = {"Username": username}
        user_scores[username][challenge] = score
    # prepare rows for excel
    headers = ["Username", "Correctness"]
    sub_headers = ["", "Challenge1", "Challenge2",
                   "challenge3", "Final Score"]
    rows = [headers, sub_headers]

    for username, scores in user_scores.items():
        row = [scores.get(header.lower(), 0) for header in sub_headers[1:-1]]
        rows.append([username]+row)

    # write to excel
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active

    # clear existing data
    sheet.delete_rows(1, sheet.max_row)

    # add formula for final score from row 2 and below
    last_row = len(rows)
    for row in range(3, last_row+1):  # Start from row 3 (E3) to last row
        sheet[f'E{row}'] = f'=SUM(B{row}, C{row}, D{row})'

    # write headers and sub-headers
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    # merge cells for correctness headers
    correctness_col_start = 2  # Column B
    correctness_col_end = len(sub_headers)
    sheet.merge_cells(start_row=1, start_column=correctness_col_start,
                      end_row=1, end_column=correctness_col_end)
    sheet.cell(row=1, column=correctness_col_start, value="Correctness")

    # save the workbook
    workbook.save(excel_file)
    logging.info(f"Result saved to {excel_file}.")


def load_challenge_config(config_file="challenges.json"):
    # ... (same as original) ...
    """
    Load challenge configurations from a JSON file.
    """
    try:
        with open(config_file, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        logging.error(f"Configuration file {config_file} not found!")
        return {}
    except json.JSONDecodeError as jde:
        logging.error(f"Error decoding JSON from {config_file}: {jde}")


def run_test_for_all_users():
    challenges = load_challenge_config()
    root_dir = os.path.dirname(os.path.abspath(__file__))
    solution_paths = defaultdict(list)

    for dirpath, _, filenames in os.walk(root_dir):
        rel_path = os.path.relpath(dirpath, root_dir)
        parts = rel_path.split(os.path.sep)
        if len(parts) == 3 and parts[2].startswith('challenge'):
            language = parts[0]
            solution_file = {
                'python': 'Solution.py',
                'java': 'Solution.java',
                'c': 'Solution.c',
                'cpp': 'Solution.cpp',
                'js': 'Solution.js'
            }.get(language)
            if solution_file and solution_file in filenames:
                challenge = parts[2]
                solution_paths[challenge].append((
                    os.path.join(dirpath, solution_file), language))
    all_results = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = []
        for challenge, paths in solution_paths.items():
            if challenge not in challenges:
                continue
            config = challenges[challenge]
            futures.append(executor.submit(
                run_test_for_challenge,
                challenge,
                paths,
                config["function_name"],
                config["test_cases"],
                config["max_score"]
            ))
        for future in futures:
            all_results.extend(future.result())
    update_excel(all_results)


if __name__ == "__main__":
    run_test_for_all_users()
